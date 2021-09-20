# -*- coding: utf-8 -*-
"""
Read and manage time series and metadata within ``flux-data-qaqc`` module
"""

import configparser as cp
import numpy as np
import pandas as pd
import refet
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from .plot import Plot
from .util import Convert

class Data(Plot, Convert):
    """
    An object for interfacing ``flux-data-qaqc`` with input metadata (config)
    and time series input, it provides methods and attributes for 
    parsing, temporal analysis, visualization, and filtering data.
    
    A :obj:`Data` object is initialized from a config file (see :ref:`Setting up
    a config file`) with metadata for an eddy covariance tower or other dataset
    containing time series meterological data. It serves as a starting point in
    the Python API of the energy balance closure analysis and data validation
    routines that are provided by ``flux-data-qaqc``. 
    
    Manual pre-filtering of data based on user-defined quality is aided with
    the :meth:`Data.apply_qc_flags` method.  Weighted or non-weighted means of
    variables with multiple sensors/recordings is performed upon initialization
    if these options are declared in the config file. The ``Data`` class also
    includes the :attr:`Data.df` property which returns the time series data in
    the form of a :obj:`pandas.DataFrame` object for custom workflows. ``Data``
    inherits line and scatter plot methods from :obj:`.Plot` which allows for
    the creation of interactive visualizations of input time series data.

    Attributes:
        climate_file (pathlib.Path): Absolute path to climate input file.
        config (:obj:`configparser.ConfigParser`): Config parser instance
            created from the data within the config.ini file.
        config_file (:obj:`pathlib.Path`): Absolute path to config.ini file 
            used for initialization of :obj:`Data` instance. 
        header (:obj:`numpy.ndarray` or :obj:`pandas.DataFrame.index`): 
            Header as found in input climate file.
        elevation (float): Site elevation in meters as set in config.ini.
        inv_map (dict): Dictionary with input climate file names as keys and 
            internal names as values. May only include pairs when they differ.
        latitude (float): Site latitude in decimal degrees, set in config.
        longitude (float): Site longitude in decimal degrees, set in config.
        out_dir (pathlib.Path): Default directory to save output of 
            :meth:`QaQc.write` or :meth:`QaQc.plot` methods.
        plot_file (pathlib.Path or None): path to plot file once it is 
            created/saved by :meth:`Data.plot`. 
        site_id (str): Site ID as found in site_id config.ini entry.
        soil_var_weight_pairs (dict): Dictionary with names and weights for
            weighted averaging of soil heat flux or soil moisture variables.
        qc_var_pairs (dict): Dictionary with variable names as keys and QC
            value columns (numeric of characters) as values.
        units (dict): Dictionary with internal variable names as keys and 
            units as found in config as values.
        variables (dict): Dictionary with internal names for variables as keys 
            and names as found in the input data as values.
        variable_names_dict (dict): Dictionary with internal variable names
            as keys and keys in config.ini file as values.
        xl_parser (str or None): engine for reading excel files with Pandas. If
            :obj:`None` use 'openpyxl'.

    """

    # maps internal names to config names for variables
    # updated with user's column names in instance attribute "variables"
    variable_names_dict = {
        'date' : 'datestring_col',
        'Rn' : 'net_radiation_col',
        'G' : 'ground_flux_col',
        'LE' : 'latent_heat_flux_col',
        'LE_user_corr' : 'latent_heat_flux_corrected_col',
        'H' : 'sensible_heat_flux_col',
        'H_user_corr' : 'sensible_heat_flux_corrected_col',
        'sw_in' : 'shortwave_in_col',
        'sw_out' : 'shortwave_out_col',
        'sw_pot' : 'shortwave_pot_col',
        'lw_in' : 'longwave_in_col',
        'lw_out' : 'longwave_out_col',
        'rh' : 'rel_humidity_col',
        'vp' : 'vap_press_col',
        'vpd' : 'vap_press_def_col',
        't_avg' : 'avg_temp_col',
        'ppt' : 'precip_col',
        'wd' : 'wind_dir_col',
        'ws' : 'wind_spd_col'
    }

    def __init__(self, config):

        self.config_file = Path(config).absolute()
        self.config = self._load_config(self.config_file)
        self.variables = self._get_config_vars()
        self.units = self._get_config_units()
        self.na_val = self.config.get('METADATA', 'missing_data_value', fallback=None)
        # try to parse na_val as numeric 
        try:
            self.na_val = float(self.na_val)
        except:
            pass
        self.elevation = float(self.config.get('METADATA', 'station_elevation'))
        self.latitude = float(self.config.get('METADATA', 'station_latitude'))
        self.longitude = float(self.config.get('METADATA', 'station_longitude'))
        self.climate_file = self._get_climate_file()
        self.xl_parser = 'openpyxl'
        self.header = self._get_header(self.climate_file)
        self.soil_var_weight_pairs = self._get_soil_var_avg_weights()
        self.qc_var_pairs = self._get_qc_flags()
        self.site_id = self.config.get('METADATA', 'site_id')
        # get optionally defined QC thresholds (numeric) or flags (string)
        # for filtering bad data
        if 'qc_threshold' in dict(self.config.items('METADATA')):
            self.qc_threshold=float(self.config.get('METADATA','qc_threshold'))
        else:
            self.qc_threshold = None
        # optional comma separated string of QC flags
        if 'qc_flag' in dict(self.config.items('METADATA')):
            self.qc_flag = [
                el.strip() for el in self.config.get(
                    'METADATA','qc_flag'
                ).split(',')
            ]
        else:
            self.qc_flag = None
        # output dir will be in directory of config file
        self.out_dir = self.config_file.parent / 'output'
        self._df = None
        self.plot_file = None


    def hourly_ASCE_refET(self, reference='short', anemometer_height=None):
        """
        Calculate hourly ASCE standardized short (ETo) or tall (ETr) reference 
        ET from input data and wind measurement height.

        If input data's time frequency is < hourly the input data will be 
        resampled to hourly and the output reference ET time series will be 
        returned as a datetime :obj:`pandas.Series` object, if the input data
        is already hourly then the resulting time series will automatically
        be merged into the :attr:`.Data.df` dataframe named "ASCE_ETo" or 
        "ASCE_ETr" respectively.

        Keyword Arguments:
            reference (str): default "short", calculate tall or short ASCE
                reference ET.
            anemometer_height (float or None): wind measurement height in meters
                , default :obj:`None`. If :obj:`None` then look for the
                "anemometer_height" entry in the **METADATA** section of the
                config.ini, if not there then print a warning and use 2 meters.

        Returns:
            :obj:`None` or :obj:`pandas.Series`

        Hint:
            The input variables needed to run this method are: vapor pressure,
            wind speed, incoming shortwave radiation, and average air 
            temperature. If vapor pressure deficit and average air temperature
            exist, the actual vapor pressure will automatically be calculated.
        """

        self.df.head(); # creates vp/vpd
        df = self.df.rename(columns=self.inv_map)

        req_vars = ['vp', 'ws', 'sw_in', 't_avg']

        if not set(req_vars).issubset(df.columns):
            print('Missing one or more required variables, cannot compute')
            return

        second_day = df.index.date[2]
        third_day = second_day + pd.Timedelta(1, unit='D')
        n_samples_per_day = len(df.loc[str(second_day)].index) 

        if n_samples_per_day < 24:
            print('Temporal frequency greater than hourly, not downsampling.')
            return

        if anemometer_height is None:
            anemometer_height = self.config.get(
                'METADATA', 'anemometer_height', fallback=None
            )
            if anemometer_height is None:
                print(
                    'WARNING: anemometer height was not given and not found in '
                    'the config files metadata, proceeding with height of 2 m'
                )
                anemometer_height = 2

        for v, u in self.units.items():
            # only converting variables needed for ref ET
            if not v in req_vars:
                continue
            if not v in Convert.required_units.keys():
                # variable is not required to have any particular unit, skip
                continue
            elif not u in Convert.allowable_units[v]:
                print('ERROR: {} units are not recognizable for var: {}\n'
                    'allowable input units are: {}\nNot converting.'.format(
                        u, v, ','.join(Convert.allowable_units[v])
                    )
                )
            elif not u == Convert.required_units[v]:
                # do conversion, update units
                # pass variable, initial unit, unit to be converted to, df
                df = Convert.convert(v, u, Convert.required_units[v], df)
                self.units[v] = Convert.required_units[v]

        # RefET will convert to MJ-m2-hr
        input_units = {
            'rs': 'w/m2'
        }

        if n_samples_per_day == 24:
            length = len(df.t_avg)
            tmean = df.t_avg
            rs = df.sw_in
            ea = df.vp
            uz = df.ws
            zw = anemometer_height
            lat = np.full(length, self.latitude)
            lon = np.full(length, self.longitude)
            doy = df.index.dayofyear
            elev = np.full(length, self.elevation)
            time = df.index.hour

        elif n_samples_per_day > 24:
            print(
                'Resampling ASCE reference ET input variables to hourly means'
            )
            tmean = df.t_avg.resample('H').mean()
            length = len(tmean)
            rs = df.sw_in.resample('H').mean()
            ea = df.vp.resample('H').mean()
            uz = df.ws.resample('H').mean()
            zw = anemometer_height
            lat = np.full(length, self.latitude)
            lon = np.full(length, self.longitude)
            doy = tmean.index.dayofyear
            elev = np.full(length, self.elevation)
            time = tmean.index.hour

        REF = refet.Hourly(
            tmean,
            ea,
            rs,
            uz,
            zw,
            elev,
            lat,
            lon,
            doy,
            time,
            method='asce',
            input_units=input_units,
        )

        if reference == 'short':
            ret = REF.eto()
            name = 'ASCE_ETo'
        elif reference == 'tall':
            ret = REF.etr()
            name = 'ASCE_ETr'

        if n_samples_per_day == 24:
            # can add directly into Data.df
            df[name] = ret
            self._df = df.rename(columns=self.variables)
            self.variables[name] = name
            self.units[name] = 'mm'

        else:
            print(
                'WARNING: cannot merge {} into the dataframe because the '
                'input temporal frequency is < hourly, returning it as an '
                'hourly, datetime-index Pandas Series'.format(name)
            )
            return pd.Series(index=tmean.index, data=ret, name=name)


    def _calc_rn(self, df):
        """
        If short and longwave radiation inputs exist but Rn is not given
        calculate it.
        """

        df = df.rename(columns=self.inv_map)

        # if Rn exists skip 
        if 'Rn' in df.columns and not df.Rn.isna().all():
            return

        rad_vars = ['sw_in', 'sw_out', 'lw_in', 'lw_out']

        has_rad_vars = set(rad_vars).issubset(df.columns)

        for v in rad_vars:
            u = self.units.get(v)
            if u:
                self.units[v] = u = u.lower()

            if u and not u in Data.allowable_units[v]:
                print('ERROR: {} units are not recognizable for var: {}\n'
                    'allowable input units are: {}\nNot converting.'.format(
                        u, v, ','.join(Data.allowable_units[v])
                    )
                )
            elif u and not u == Data.required_units[v]:
                # do conversion, update units
                df = Convert.convert(v, u, Data.required_units[v], df)
                self.units[v] = Data.required_units[v]

        units_correct = (
            self.units.get('sw_in') == 'w/m2'  and \
            self.units.get('sw_out') == 'w/m2' and \
            self.units.get('lw_in') == 'w/m2'  and \
            self.units.get('lw_out') == 'w/m2'
        )

        if has_rad_vars and units_correct:
            print('Calculating net radiation from components.')
            df['Rn'] = df.sw_in + df.lw_in - df.sw_out - df.lw_out
            self.variables['Rn'] = 'Rn'
            self.units['Rn'] = 'w/m2'

        self._df = df


    def _calc_vpd_or_vp(self, df):
        """
        Based on ASCE standardized ref et eqn. 37, air temperature must be in 
        celcius and actual vapor pressure in kPa.

        Can also calculate VP from VPD and air temperature.
        """
        df = df.rename(columns=self.inv_map)

        # make sure day intervals are hourly or less if not skip
        second_day = df.index.date[2]
        third_day = second_day + pd.Timedelta(1, unit='D')
        # both days start at 00:00:00, don't duplicate
        times_in_day = len(df.loc[str(third_day)].index) 
        if times_in_day < 24:
            print('Temporal frequency of data > hourly cannot calculate VP/VPD')
            return

        for v in ['vp', 'vpd', 't_avg']:
            u = self.units.get(v)
            if u:
                self.units[v] = u = u.lower()

            if u and not u in Data.allowable_units[v]:
                print('ERROR: {} units are not recognizable for var: {}\n'
                    'allowable input units are: {}\nNot converting.'.format(
                        u, v, ','.join(Data.allowable_units[v])
                    )
                )
            elif u and not u == Data.required_units[v]:
                # do conversion, update units
                # pass variable, initial unit, unit to be converted to, df
                df = Convert.convert(v, u, Data.required_units[v], df)
                self.units[v] = Data.required_units[v]

        # calculate vpd from actual vapor pressure and temp
        # check if needed variables exist and units are correct
        has_vpd_vars = set(['vp','t_avg']).issubset(df.columns)
        units_correct = (
            self.units.get('vp') == 'kpa' and self.units.get('t_avg') == 'c'
        )
        if has_vpd_vars and units_correct:
            print(
                'Calculating vapor pressure deficit from vapor pressure and '
                'air temperature'
            )
            # saturation vapor pressure (es)
            es = 0.6108 * np.exp(17.27 * df.t_avg / (df.t_avg + 237.3))
            df['vpd'] = es - df.vp
            df['es'] = es
            self.variables['vpd'] = 'vpd'
            self.units['vpd'] = 'kpa'
            self.variables['es'] = 'es'
            self.units['es'] = 'kpa'

        # same calc actual vapor pressure from vapor pressure deficit and temp
        has_vp_vars = set(['vpd','t_avg']).issubset(df.columns)
        units_correct = (
            self.units.get('vpd') == 'kpa' and self.units.get('t_avg') == 'c'
        )

        if has_vp_vars and units_correct:
            print(
                'Calculating vapor pressure from vapor pressure deficit and '
                'air temperature'
            )
            # saturation vapor pressure (es)
            es = 0.6108 * np.exp(17.27 * df.t_avg / (df.t_avg + 237.3))
            df['vp'] = es - df.vpd
            df['es'] = es
            self.variables['vp'] = 'vp'
            self.units['vp'] = 'kpa'
            self.variables['es'] = 'es'
            self.units['es'] = 'kpa'

        if not 'rh' in self.variables and {'vp','es'}.issubset(self.variables):
            if not self.units.get('vp') == 'kpa': pass
            else:
                print(
                    'Calculating relative humidity from actual and saturation '
                    'vapor pressure and air temperature'
                )
                df['rh'] = 100 * (df.vp / df.es)
                self.variables['rh'] = 'rh'
                self.units['rh'] = '%'
        
        if 'vp' in self.variables and self.units.get('vp') == 'kpa':
            print(
                'Calculating dew point temperature from vapor pressure'
            )
            df['t_dew'] = (-1 / ((np.log(df.vp/.611) / 5423) - (1/273)))-273.15
            self.variables['t_dew'] = 't_dew'
            self.units['t_dew'] = 'c'

        self._df = df


    def plot(self, ncols=1, output_type='save', out_file=None, suptitle='', 
            plot_width=1000, plot_height=450, sizing_mode='scale_both', 
            merge_tools=False, link_x=True, **kwargs):
        """
        Creates a series of interactive diagnostic line and scatter 
        plots of input data in whichever temporal frequency it is in.

        The main interactive features of the plots include: pan, selection and
        scrol zoom, hover tool that shows paired variable values including date,
        and linked x-axes that pan/zoom togehter for daily and monthly time 
        series plots.
 
        It is possible to change the format of the output plots including
        adjusting the dimensions of subplots, defining the number of columns of
        subplots, setting a super title that accepts HTML, and other options.
        If variables are not present for plots they will not be created and a
        warning message will be printed. There are two options for output: open
        a temporary file for viewing or saving a copy to :attr:`QaQc.out_dir`.
        
        A list of all potential time series plots created: 
        
        * energy balance components 
        * radiation components 
        * multiple soil heat flux measurements
        * air temperature
        * vapor pressure and vapor pressure deficit
        * wind speed
        * precipitation 
        * latent energy
        * multiple soil moisture measurements


        Keyword Arguments:
            ncols (int): default 1. Number of columns of subplots.
            output_type (str): default "save". How to output plots, "save", 
                "show" in browser, "notebook" for Jupyter Notebooks, 
                "return_figs" to return a list of Bokeh 
                :obj:`bokeh.plotting.figure.Figure`s, or "return_grid" to 
                return the :obj:`bokeh.layouts.gridplot`.
            out_file (str or None): default :obj:`None`. Path to save output 
                file, if :obj:`None` save output to :attr:`Data.out_dir` with 
                the name [site_id]_input_plots.html where [site_id] is 
                :attr:`Data.site_id`.
            suptitle (str or None): default :obj:`None`. Super title to go 
                above plots, accepts HTML/CSS syntax.
            plot_width (int): default 1000. Width of subplots in pixels.
            plot_height (int): default 450. Height of subplots in pixels, note 
                for subplots the height will be forced as the same as 
                ``plot_width``.
            sizing_mode (str): default "scale_both". Bokeh option to scale
                dimensions of :obj:`bokeh.layouts.gridplot`.
            merge_tools (bool): default False. Merges all subplots toolbars into
                a single location if True.
            link_x (bool): default True. If True link x axes of daily time 
                series plots and monthly time series plots so that when zooming 
                or panning on one plot they all zoom accordingly, the axes will 
                also be of the same length.

        Example:
            
            Starting from a correctly formatted config.ini and climate time
            series file, this example shows how to read in the data and produce
            a series of plots of input data as it is found in the input data
            file (unlike :meth:`.QaQc.plot` which produces plots at daily and
            monthly temporal frequency). This example also shows how to display
            a title at the top of plot with the site's location and site ID.

            >>> from fluxdataqaqc import Data
            >>> d = Data('path/to/config.ini')
            >>> # create plot title from site ID and location in N. America
            >>> title = "<b>Site:</b> {}; <b>Lat:</b> {}N; <b>Long:</b> {}W".format(
            >>>     q.site_id, q.latitude, q.longitude
            >>> )
            >>> q.plot(
            >>>     ncols=2, output_type='show', plot_width=500, suptitle=title
            >>> )

            Note, we specified the width of plots to be smaller than default
            because we want both columns of subplots to be viewable on the 
            screen.

        Tip:
            To reset all subplots at once, refresh the page with your web
            browser.

        Note:
            Additional keyword arguments that are recognized by
            :obj:`bokeh.layouts.gridplot` are also accepted by
            :meth:`Data.plot`.

        See Also:
            :meth:`.QaQc.plot`
        """

        # handle file save for accessing from instance variable
        if out_file is None and output_type == 'save':
            out_file = Path(self.out_dir)/'{}_input_plots.html'.format(
                self.site_id
            )
            out_dir = out_file.parent
            if not out_dir.is_dir():
                out_dir.mkdir(parents=True, exist_ok=True)
        # to allow making any subdir that does not yet exist
        # if out_file is to a non-existent directory create parents
        elif out_file is not None and output_type == 'save':
            out_dir = Path(out_file).parent
            if not out_dir.is_dir():
                out_dir.mkdir(parents=True, exist_ok=True)
        
        # create aggregrated plot structure from fluxdataqaqc.Plot._plot() 
        ret = self._plot(
            self, ncols=ncols, output_type=output_type, out_file=out_file,
            suptitle=suptitle, plot_width=plot_width, plot_height=plot_height,
            sizing_mode=sizing_mode, merge_tools=merge_tools, link_x=link_x,
            **kwargs
        )

        self.plot_file = out_file

        if ret:
            return ret

    def _load_config(self, config_file):
        if not config_file.is_file():
            raise FileNotFoundError('ERROR: config file not found')
        config = cp.ConfigParser(interpolation=None)
        config.read(config_file)
        return config

    def _get_climate_file(self):
        """
        Read config file and return absolute path to climate time series file
        """
        file_path = self.config['METADATA']['climate_file_path']
        climate_file = self.config_file.parent.joinpath(file_path)
        
        if not climate_file.is_file():
            err_msg = 'ERROR: climate file:{} does not exist'.format(
                climate_file)
            raise FileNotFoundError(err_msg)
        
        return climate_file

    def _get_header(self, climate_file):
        """
        Read only top line of climate time series file return header names.
        """
        if climate_file.suffix in ('.xlsx', '.xls'):
            try: # using xlrd
                workbook = pd.ExcelFile(climate_file)
                rows = workbook.book.sheet_by_index(0).nrows
                header = pd.read_excel(workbook, skipfooter = (rows - 1))
                header = header.columns
            except: # fallback openpyxl- slower
                wb = load_workbook(climate_file, enumerate)
                sheet = wb.worksheets[0]
                header = sheet._shared_strings
                self.xl_parser='openpyxl'
                #rows = sheet.max_row
                #header = pd.read_excel(workbook, skipfooter = (rows - 1))
                #header = header.columns
        
        else: # assume CSV
            skiprows=None
            if 'skiprows' in dict(self.config.items('METADATA')):
                val = self.config.get('METADATA','skiprows')
                if val and val.startswith('[') and val.endswith(']'):
                    skiprows = val.replace(' ','')
                    skiprows = [
                        int(el) for el in skiprows.strip('][').split(',')
                    ] 
                else:
                    skiprows = int(val)

            header = pd.read_csv(
                climate_file, skiprows=skiprows, nrows=0
            ).columns

        return header

    def _get_soil_var_avg_weights(self):
        """
        Read config and variable attribute to match user names for multiple
        soil moisture/heat flux variables to user defined weights used for
        calculating weighted/non-weighted average values in `Data.df`
        """
        soil_var_weight_pairs = {}
        all_keys = dict(self.config.items('DATA')).keys()

        for k,v in dict(self.config.items('DATA')).items():
            weight_name = '{}_weight'.format(k)
            var_name = self.variables.get(k)
            if k in self.variables.keys() and weight_name in all_keys:
                weight = self.config.get('DATA', weight_name)
                tmp = {'name': var_name, 'weight' : weight}
                soil_var_weight_pairs[k] = tmp
                #print(k,v, var_name, weight)
            # update dict with weights if not specified in config, normalized
            # later in df
            elif (k.startswith('g_') or k.startswith('theta_')) and not \
                    weight_name in all_keys and not k.endswith('_units') and \
                    not k.endswith('_weight'):
                tmp = {'name': var_name, 'weight' : 1}
                soil_var_weight_pairs[k] = tmp

        return soil_var_weight_pairs

    def _get_config_vars(self):
        """
        Read config data section and get names of all variables, pair with
        internal variable names in a dictionary.

        Also parses config file for optionally added multiple soil heat flux
        and soil moisture variables if given following the naming convention
        explained in ``flux-data-qaqc``.

        Arguments:
            None

        Returns:
            variables (dict): dictionary with internal climate variable names as keys and user's names as found in config as values. 
        """

        variables = {}
        # get all variables found in Data.variable_names_dict
        for k, v in Data.variable_names_dict.items():
            if self.config.has_option('DATA', v):
                variables[k] = self.config.get('DATA', v)

        # get multiple G flux/soil moisture variables 
        all_keys = dict(self.config.items('DATA')).keys()
        # should be named as 'g_ or 'theta_ what comes after not strict yet
        added_g_keys = []
        added_theta_keys = []
        for k in all_keys:
            if k.startswith('g_') and not k.endswith('_units') \
                    and not k.endswith('_weight'):
                added_g_keys.append(k)
            if k.startswith('theta_') and not k.endswith('_units') \
                    and not k.endswith('_weight'):
                added_theta_keys.append(k)
                
        if added_g_keys:
            for el in added_g_keys:
                variables[el] = self.config.get('DATA', el)
        if added_theta_keys:
            for el in added_theta_keys:
                variables[el] = self.config.get('DATA', el)
        
        # update all None values in config with 'na' in variables dict
        for k, v in variables.items():
            if v is None:
                variables[k] = 'na'

        return variables
    
    def _get_config_units(self):
        """
        Get units from config file, pair with var names and store in dictionary.

        Keys are `flux-data-qaqc` variable names, i.e. the same names used 
        by :attr:`Data.variables` and values are strings assigned in the config
        file. 

        Arguments:
            None

        Returns:
            units (dict): dictionary with `flux-data-qaqc` variable names as keys and user's units for each as values.

        Note:
            Parsing of correct units and conversion if needed is performed
            in the :obj:`.QaQc` class. Also, if units are not given
            in the config file a warning message is printed and the units are
            not included and thus will either need to be manually added later
            e.g. in Python by adding to :attr:`Data.units` or by adding them
            to the config and recreating a :obj:`Data` object otherwise the 
            units will remain unknown and not be able to be later converted.
        """
        no_unit_vars = ('datestring_col', 'year_col', 'month_col', 'day_col')
        config_dict = dict(self.config.items('DATA'))
        # dictionary that maps config unit keys to units
        units_config = {
            v.replace('_col', '_units'): None for k,v in 
            self.variable_names_dict.items() if (
                not v in no_unit_vars and k in self.variables
            )
        }
        # add user multiple g or soil moisture var units config names
        for k,v in self.variables.items():
            # if multiple g uses same var assigned to ground_flux_col units the 
            # added G var will not be included/duplicated
            if k.startswith('g_') and not self.variables[k] == 'G':
                units_config['{}_units'.format(k)] = None
            if k.startswith('theta_'):
                units_config['{}_units'.format(k)] = None

        for k in units_config: 
            if k in config_dict:
                units_config[k] = config_dict[k]
            # user corrected versions of LE, H, etc. are in variables_name_dict
            # but may not be included in the input config/data
            elif '_corrected' in k:
                continue
            else:
                print(
                    'WARNING: units for var {} missing from the config file'\
                        .format(k.replace('_units',''))
                )
                
        inv_names = {v:k for k,v in self.variable_names_dict.items()}

        # dictionary that maps fluxdataqaqc var names to units
        units = dict()
        for k in units_config:
            k_col = k.replace('_units', '_col')
            if k_col in self.variable_names_dict.values():
                units[inv_names[k_col]] = units_config.get(k)
            else:
                # for multiple g and theta remove _units suffix
                name = k.replace('_units','')
                units[name] = units_config.get(k)

        return units

    def _get_qc_flags(self):
        """
        Process any existing QC flags for variables in config, also add
        key,val to variables attribute if QC flags are not specified in config
        and follow the naming convention [var_name]_QC in the header of the 
        data file.
        """
        qc_var_pairs = {}
        tmp = {}

        no_qc_vars = ('datestring_col')
        # dictionary that maps config QC values to keys for main variables
        # other variables like multiple g or theta (with unknown names) are
        # search for in loop below
        qc_config = {
            v.replace('_col', '_qc'): k for k,v in 
            self.variable_names_dict.items() if not v in no_qc_vars
        }
        # first look if specified in config
        for k,v in self.config.items('DATA'):
            if k in qc_config:
                # internal name for the variable (e.g. LE or Rn)
                var_name = qc_config[k]
                user_var_name = self.variables.get(var_name)
            # keys are internal names for multiple G and theta
            elif k.startswith(('g_','theta_')) and k.endswith('_qc'):
                var_name = k 
            # key is not for a QC flag header name...
            else:
                continue
            if not v in self.header:
                print('WARNING: {} quality control name specified in the config'
                    ' file for variable: {} does not exist in the input file, '
                    'it will not be used.'.format(v, self.variables[var_name])
                )
                continue
            internal_name = '{}_qc_flag'.format(var_name)
            tmp[internal_name] = v
            qc_var_pairs[user_var_name] = v

        # also look in header, currently always gets these as well, may change 
        # find vairable names that have a matching name with '_QC' suffix
        # if found here but different name in config use the name in the config 
        for k,v in self.variables.items():
            qc_var = '{}_QC'.format(v)
            if qc_var in self.header:
                internal_name = '{}_qc_flag'.format(k)
                if internal_name in tmp:
                    continue
                tmp[internal_name] = qc_var
                qc_var_pairs[v] = qc_var
               
        self.variables.update(tmp)
        return qc_var_pairs

    def apply_qc_flags(self, threshold=None, flag=None):
        """
        Apply user-provided QC values or flags for climate variables to filter
        poor-quality data by converting them to null values, updates 
        :attr:`Data.df`. 
        
        Specifically where the QC value is < `threshold` change the variables
        value for that date-time to null. The other option is to use a column
        of flags, e.g. 'x' for data values to be filtered out. The threshold
        value or flag may be specified in the config file's **METADATA**
        section otherwise they should be assigned as keyword arguments here.

        Specification of which QC (flag or numeric threshold) columns should be
        applied to which variables is set in the **DATA** section of the config
        file. For datasets with QC value columns with names identical to the
        variable they correspond to with the suffix "_QC" the QC column names
        for each variable do not need to be specified in the config file. 
        
        Keyword Arguments:
            threshold (float): default :obj:`None`. Threshold for QC values, if 
                flag is below threshold replace that variables value with null.
            flag (str, list, or tuple): default :obj:`None`. Character flag 
                signifying data to filter out. Can be list or tuple of multiple
                flags.

        Returns:
            :obj:`None`

        Example:

            If the input time series file has a column with numeric quality 
            values named "LE_QC" which signify the data quality for latent
            energy measurements, then in the config.ini file's **DATA** section
            the following must be specified::

                [DATA]
                latent_heat_flux_qc = LE_QC
                ...

            Now you must specify the threshold of this column in which to filter
            out when using :meth:`Data.apply_qc_flags`. For example if you want
            to remove all data entries of latent energy where the "LE_QC" value
            is below 5, then the threshold value would be 5. The threshold
            can either be set in the config file or passed as an argument. If it
            is set in the config file, i.e.::

                [METADATA]
                qc_threshold = 0.5

            Then you would cimply call the method and this threshold would be
            applied to all *QC* columns specified in the config file,

            >>> from fluxdataqaqc import Data
            >>> d = Data('path/to/config.ini')
            >>> d.apply_qc_flags()

            Alternatively, if the threshold is not defined in the config file or
            if you would like to use a different value then pass it in,

            >>> d.apply_qc_flags(threshold=2.5)

            Lastly, this method also can filter out based on a single or list
            of character flags, e.g. "x" or "bad" gievn that the column
            containing these is specified in the config file for whichever
            variable they are to be applied to. For example, if a flag column
            contains multiple flags signifying different data quality control
            info and two in particular signify poor quality data, say "b" and
            "a", then apply them either in the config file::


                [METADATA]
                qc_flag = b,a

            Of within Python

            >>> d.apply_qc_flags(flag=['b', 'a'])

            For more explanation and examples see the "Configuration
            Options" section of the online documentation.

        """
        # if QC threshold or flags not passed use values from config if exist
        if not threshold:
            threshold = self.qc_threshold
        if not flag:
            flag = self.qc_flag
        # load dataframe if not yet accessed
        df = self.df
        # infer each columns datatype to avoid applying thresholds to strings
        infer_type = lambda x: pd.api.types.infer_dtype(x, skipna=True)
        df_types = pd.DataFrame(df.apply(
            pd.api.types.infer_dtype, axis=0, skipna=True
            )
        ).rename(columns={'index': 'index', 0: 'type'})
        # loop over each variable that has a provided qc value and set nulls
        # where qc value < threshold, qc_var_pairs maps var names to qc names
        if threshold:
            for var, qc in self.qc_var_pairs.items():
                if df_types.loc[qc,'type'] != 'string':
                    df.loc[
                        (df[qc] < threshold) & (df[qc].notnull()) , var
                    ] = np.nan
        # set values to null where flag is a certain string
        if flag:
            if isinstance(flag, str):
                for var, qc in self.qc_var_pairs.items():
                    if df_types.loc[qc,'type'] == 'string':
                        df.loc[
                            (df[qc] == flag) & (df[qc].notnull()) , var
                        ] = np.nan
            # apply multiple character flags
            elif isinstance(flag, (list,tuple)):
                for f in flag:
                    for var, qc in self.qc_var_pairs.items():
                        if df_types.loc[qc,'type'] == 'string':
                            df.loc[
                                (df[qc] == f) & (df[qc].notnull()) , var
                            ] = np.nan

        self._df = df

    @property
    def df(self):
        """
        Pull variables out of the config and climate time series files load 
        them into a datetime-indexed :obj:`pandas.DataFrame`. 

        Metadata about input time series file format: "missing_data_value",
        "skiprows", and "date_parser" are utilized when first loading the
        ``df`` into memory. Also, weighted and non-weighted averaging of
        multiple measurements of the same climatic variable occurs on the first
        call of :attr:`Data.df`, if these options are declared in the config
        file. For more details and example uses of these config options please
        see the "Configuration Options" section of the online documentation.

        Returns:
            df (:obj:`pandas.DataFrame`)

        Examples:

            You can utilize the df property as with any :obj:`pandas.DataFrame`
            object. However, if you would like to make changes to the data you
            must first make a copy, then make the changes and then reassign it
            to :attr:`Data.df`, e.g. if you wanted to add 5 degrees to air temp.

            >>> from fluxdataqaqc import Data
            >>> d = Data('path_to_config.ini')
            >>> df = d.df.copy()
            >>> df['air_temp_column'] = df['air_temp_column'] + 5
            >>> d.df = df
            
            The functionality shown above allows for user-controlled
            preprocessing and modification of any time series data in the
            initial dataset. It also allows for adding new columns but if
            they are variables used by ``flux-data-qaqc`` e.g. Rn or other
            energy balance variables, be sure to also update
            :attr:`Data.variables` and :attr:`Data.units` with the appropriate
            entries. New or modified values will be used in any further
            analysis/ploting routines within ``flux-data-qaqc``.

            By default the names of variables as found within input data are
            retained in :attr:`QaQc.df`, however you can use the naming scheme
            as ``flux-data-qaqc`` which can be viewed in 
            :attr:`Data.variable_names_dict` by using the the 
            :attr:`Data.inv_map` dictionary which maps names from user-defined
            to internal names (as opposed to :attr:`Data.variables`) which
            maps from internal names to user-defined. For example if your 
            input data had the following names for LE, H, Rn, and G set in your
            config::

                [DATA]
                net_radiation_col = Net radiation, W/m2
                ground_flux_col = Soil-heat flux, W/m2
                latent_heat_flux_col = Latent-heat flux, W/m2
                sensible_heat_flux_col = Sensible-heat flux, W/m2

            Then the :attr:`Data.df` will utilize the same names, e.g. 

            >>> # d is a Data instance
            >>> d.df.head()

            produces:

            ============== =================== ====================== ======================== ==================== 
            date           Net radiation, W/m2 Latent-heat flux, W/m2 Sensible-heat flux, W/m2 Soil-heat flux, W/m2 
            ============== =================== ====================== ======================== ==================== 
            10/1/2009 0:00 -54.02421778        0.70761                0.95511                  -40.42365926         
            10/1/2009 0:30 -51.07744708        0.04837                -1.24935                 -33.35383253         
            10/1/2009 1:00 -50.99438925        0.68862                1.91101                  -43.17900525         
            10/1/2009 1:30 -51.35032377        -1.85829               -15.4944                 -40.86201497         
            10/1/2009 2:00 -51.06604228        -1.80485               -19.1357                 -39.80936855         
            ============== =================== ====================== ======================== ====================

            Here is how you could rename your dataframe using 
            ``flux-data-qaqc`` internal names,

            >>> d.df.rename(columns=q.inv_map).head()

            ============== =================== ====================== ======================== ==================== 
            date           Rn                  LE                     H                        G 
            ============== =================== ====================== ======================== ==================== 
            10/1/2009 0:00 -54.02421778        0.70761                0.95511                  -40.42365926         
            10/1/2009 0:30 -51.07744708        0.04837                -1.24935                 -33.35383253         
            10/1/2009 1:00 -50.99438925        0.68862                1.91101                  -43.17900525         
            10/1/2009 1:30 -51.35032377        -1.85829               -15.4944                 -40.86201497         
            10/1/2009 2:00 -51.06604228        -1.80485               -19.1357                 -39.80936855         
            ============== =================== ====================== ======================== ====================

            A minor note on variable naming, if your input data variables
            use exactly the same names used by ``flux-data-qaqc``, they
            will be renamed by adding the prefix "input\_", e.g. "G" becomes
            "input_G" on the first time reading the data from disk, i.e. the
            first time accessing :attr:`Data.df`.
   

        Note:
            The temporal frequency of the input data is retained unlike the
            :attr:`.Qaqc.df` which automatically resamples time series data to 
            daily frequency.  
        """

        # avoid overwriting pre-assigned data
        if isinstance(self._df, pd.DataFrame):
            return self._df.rename(columns=self.variables)

        variables = self.variables
        # remove variable entries that were given as 'na' in config
        vars_notnull = dict((k, v) for k, v in variables.items() if v != 'na')
        self.variables = vars_notnull
        cols = list(vars_notnull.values())

        # if multiple columns were assign to a variable parse them now to
        # make sure they all exist, below calculate mean and rename to _mean
        # if no var_name_delim in metadata then assume one column per var
        delim = None
        if 'var_name_delim' in dict(self.config.items('METADATA')):
            delim = self.config.get('METADATA','var_name_delim')
            cols = [x.split(delim) if delim in x else x for x in cols]
            cols_flat = []
            for el in cols:
                if isinstance(el,list):
                    cols_flat += el
                else:
                    cols_flat.append(el)
            cols = cols_flat

        missing_cols = None

        if not set(cols).issubset(self.header):
            missing_cols = set(cols) - set(self.header)
            err_msg = ('WARNING: the following config variables are missing '
                'in the input climate file:\n{}\nThey will be filled with '
                'NaN values'.format(' '.join(missing_cols)))
            print(err_msg)
            cols = set(cols).intersection(self.header)

        kwargs = {}
        if 'missing_data_value' in dict(self.config.items('METADATA')):
            self.na_val = self.config.get('METADATA', 'missing_data_value')
            # try to parse na_val as numeric 
            try:
                self.na_val = float(self.na_val)
            except:
                pass
            kwargs['na_values'] =  [self.na_val]

        if 'skiprows' in dict(self.config.items('METADATA')):
            val = self.config.get('METADATA','skiprows')
            if val and val.startswith('[') and val.endswith(']'):
                skiprows = val.replace(' ','')
                kwargs['skiprows'] = [
                    int(el) for el in skiprows.strip('][').split(',')
                ] 
                 
            else:
                kwargs['skiprows'] = int(val)

        if 'date_parser' in dict(self.config.items('METADATA')):
            date_parse_str = self.config.get('METADATA','date_parser')
            date_parser = lambda x: datetime.strptime(x, date_parse_str)
            kwargs['date_parser'] = date_parser
        if 'load_all_vars' in dict(self.config.items('METADATA')):
            # if this option is listed (with any value) read all columns into df
            cols = self.header

        # load data file depending on file format
        if self.climate_file.suffix in ('.xlsx', '.xls'):
            # find indices of headers we want, only read those, excel needs ints
            ix=[i for i, e in enumerate(self.header) if e in set(cols)]
            df = pd.read_excel(
                self.climate_file,
                parse_dates = [variables.get('date')],
                usecols = ix,
                engine = self.xl_parser,
                **kwargs
            )
        else:
            df = pd.read_csv(
                self.climate_file,
                parse_dates = [variables.get('date')],
                usecols = cols,
                **kwargs
            )

        if 'missing_data_value' in dict(self.config.items('METADATA')):
            # force na_val because sometimes with read_excel it doesn't work...
            df[df == self.na_val] = np.nan

        if missing_cols:
            df = df.reindex(columns=list(cols)+list(missing_cols))
        
        def calc_weight_avg(d, pref, df):
            """
            Helper function to reduce redundant code for calculating weighted
            average currently only for multiple soil heat flux and moisture 
            variables.

            d is soil_var_weight_pairs dict
            pref is variable prefix str, i.e. g_ or theta_
            df is the dataframe 
            """
            # list of multiple variables with prefix
            if d:
                vs = [d.get(el) for el in d if el.startswith(pref)]
            else:
                vs = []
            # soil heat flux weighted average
            if len(vs) > 1: # if 1 or less no average
                weights = [float(el.get('weight')) for el in vs]
                total_weights = np.sum(weights)
                # if multiple Gs specified and same num multiple Gs specifed
                # as the main G var and no weights assigned do not duplicate
                # mean that is calculated below from comma separated list
                if delim and pref == 'g_' and self.variables.get('G') is not \
                        None and delim in self.variables.get('G'):
                    n_Gs = len(self.variables.get('G').split(delim))
                    if len(weights) == total_weights and len(weights) == n_Gs: 
                        return
                # same for multiple theta, this also avoids issue with multiple
                # null columns and weighting the remaining columns less...
                if pref == 'theta_' and len(weights) == total_weights:
                    # update variables
                    var_name = 'theta_mean'
                    theta_names = [el.get('name') for el in vs]
                    print('Calculating mean for var: THETA from '
                        'columns: {}'.format(theta_names)
                    )
                    tmp_df = df[theta_names].copy()
                    df[var_name] = tmp_df.mean(axis=1)
                    self.variables[var_name] = var_name
                    return
                # if weights are not normalized update them
                elif not np.isclose(total_weights, 1.0):
                    print(
                        "{} weights not given or don't sum to one, normalizing"\
                            .format(pref.replace('_',''))
                    )
                    for k,v in d.items():
                        if k.startswith(pref):
                            nw = float(v.get('weight')) / total_weights
                            d[k] = {'name': v.get('name'), 'weight': nw}
                            msg = ', '.join([
                                '{}:{:.2f}'.format(
                                    v.get('name'),float(v.get('weight'))
                                    )
                                    for k,v in d.items() if k.startswith(pref)
                                ]
                            )
                    print('Here are the new weights:\n', msg)

                # calculate average, use updated weights if calculated
                vs = [d.get(el) for el in d if el.startswith(pref)]
                tmp_df = df[[el.get('name') for el in vs]].copy()
                for pair in vs:
                    tmp_df[pair.get('name')] *= float(pair.get('weight'))

                # update variables
                key = '{}mean'.format(pref)
                val = '{}mean'.format(pref)
                if key == 'g_mean':
                    key = 'G'
                # calculate mean (sum weighted values)
                df[val] = tmp_df.sum(axis=1)
                df.loc[df[tmp_df.columns].isnull().all(1), val] = np.nan

                self.variables[key] = val
            elif len(vs) == 1:
                # only print this message if 1 weighted avg var is assigned...
                print(
                    'WARNING: Insufficient data to calculate mean for multiple '
                    '{} measurements'.format(pref.replace('_','').upper())
                )
                # remove from var name dict if G to avoid inv_map overwrite
                # i.e. only one G recording but listed twice in config 
                if pref == 'g_':
                    to_drop = [i for i in d.keys() if i.startswith('g_')][0]
                    self.variables.pop(to_drop, None)
                    self.units.pop(to_drop, None)


        # check if any of multiple soil vars time series are all null and remove
        del_keys = []
        for k, v in self.soil_var_weight_pairs.items():
            var_name = v.get('name')
            if df[var_name].isna().all():
                del_keys.append(k)
        for k in del_keys:
            self.soil_var_weight_pairs.pop(k)

        # calculate weighted average soil vars if they exist
        d = self.soil_var_weight_pairs
        calc_weight_avg(d, 'g_', df)
        calc_weight_avg(d, 'theta_', df)

        # currently calc non weighted means for vars other than G and theta
        # later may change so all vars are handled the same way, this is for
        # multiple listed (delim sep) variables for a single var in the config 
        for k,v in variables.items():
            if delim and delim in v:
                tmp_cols = v.split(delim) 
                print('Calculating mean for var: {}\n'.format(k),
                    'from columns: {}\n'.format(tmp_cols)
                )
                # calc mean and update variables dict names 
                var_name = k+'_mean'
                df[var_name] = df[tmp_cols].mean(axis=1)
                self.variables[k] = var_name

        # check each variable for multiple criteria, all null, naming, etc.
        del_keys = []
        for k,v in self.variables.items():
            if v not in df.columns:
                del_keys.append(k)
            # the all zero issue is a weird bug with pandas mean or nans that
            # may depend on a dependency of pandas called bottleneck
            elif df[v].isnull().all() or (df[v] == 0).all():
                print(
                    'WARNING: {} variable in column {} is missing all data '
                    'it will be removed'.format(k, v)
                )
                df.drop(v, axis=1, inplace=True)
                del_keys.append(k)
            # also rename input columns that may cause naming issue
            elif k in df.columns and k in Data.variable_names_dict.keys() and \
                    not k == 'date':
                new_name = 'input_{}'.format(k)
                print('WARNING: renaming column {} to {}'.format(k, new_name))
                df.rename(columns={k:new_name}, inplace=True)
                if not v == k+'_mean':
                    self.variables[k] = new_name
                if k in self.qc_var_pairs:
                    self.qc_var_pairs[new_name] = self.qc_var_pairs[k] 
                    self.qc_var_pairs.pop(k, None)

        # also remove entry from var dict to avoid issues later
        for k in del_keys:
            self.variables.pop(k, None)
            self.units.pop(k, None)

        # update renaming dict with any newly created mean variables/removed
        self.inv_map = {
            v: k for k, v in self.variables.items() if not k == v
        }

        # the only column that is always renamed is the datestring_col
        df.rename(columns={variables['date']: 'date'}, inplace=True)
        
        # date index
        df.index = df.date
        df = df[df.index.notnull()]
        df.drop('date', axis=1, inplace=True)
        self._df = df # vpd calc uses attribute
        # calc vapor pressure or vapor pressure deficit if hourly or less
        # also converts units if needed for vp, vpd, t_avg
        self._calc_vpd_or_vp(df)
        self._calc_rn(df)

        return df

    @df.setter
    def df(self, data_frame):
        if not isinstance(data_frame, pd.DataFrame):
            raise TypeError("Must assign a pandas.DataFrame object")
        self._df = data_frame

